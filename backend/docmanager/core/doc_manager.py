from collections import OrderedDict
import html
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Set
from docxtpl import DocxTemplate
from jinja2 import Environment
from xml.sax.saxutils import escape
from openpyxl import load_workbook

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="docxcompose.properties")

class DocError(Exception):
    """Base error for Doc logic."""
    pass

class TemplateNotFound(DocError):
    """Template not found or not a .docx file."""
    pass

class InvalidContext(DocError):
    """Context has wrong type/structure."""
    pass

class RenderError(DocError):
    """Error during rendering (docxtpl/jinja)."""
    pass
    
Reserved = {"meta", "options", "layout"}
Reserved_placeholders = {"PROVIDER_ADDR"}

class DocManager:
    def __init__(self, base_dir: Path):
        self.base_dir = Path(base_dir).resolve()
        
    # ---------- helpers ----------
    def _unescape(self, obj: Any) -> Any:
        if isinstance(obj, str): 
            return html.unescape(obj)
        if isinstance(obj, dict): 
            return {k: self._unescape(v) for k, v in obj.items()}
        if isinstance(obj, list): 
            return [self._unescape(v) for v in obj]
        return obj
    
    def _sanitize(self, obj: Any):
        if isinstance(obj, str):
            s = html.unescape(obj)
            return escape(s, {"'": "&apos;", '"': "&quot;"})
        if isinstance(obj, dict):
            return {k: self._sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [self._sanitize(v) for v in obj]
        return obj
            
    def _template_path(self, filename: str) -> Path:
        p = (self.base_dir / filename).resolve()
        if not (p.is_file() and p.suffix.lower() == ".docx"):
            raise TemplateNotFound(f"Template not found: {filename}")
        return p
    
    def _useReservedPlaceholders(self, ctx: Dict[str, Any]):
        if "PROVIDER" in ctx and ctx["PROVIDER"]:
            ctx["PROVIDER_ADDR"] = self._load_provider_addr("provider_addresses.xlsx", ctx["PROVIDER"])
        
    
    # ---------- reading ----------
    def get_placeholders(self, filename: str) -> Set[str]:
        tpl = DocxTemplate(self._template_path(filename).as_posix())
        return set(tpl.get_undeclared_template_variables())
    
    def _call_collect(self, filename: str, func_name: str) -> List[dict]:
        """Executes custom Jinja calls in the template like:
            {% set _ = options( [ ... ] ) %}
            {% set _ = meta( [ ... ] ) %}
            {% set _ = layout( [ ... ] ) %}
        and collects the return values (always "") in Python
        """
        collected: List[dict] = []
        
        def _collector(arg):
            collected.append(arg)
            return ""
        
        tpl = DocxTemplate(self._template_path(filename).as_posix())
        jinja_env = Environment()
        jinja_env.globals["options"] = _collector if func_name == "options" else lambda *_:""
        jinja_env.globals["meta"] = _collector if func_name == "meta" else lambda *_:""
        jinja_env.globals["layout"] = _collector if func_name == "layout" else lambda *_:""
        tpl.render({}, jinja_env)
        
        return collected
    
    def get_options(self, filename: str) -> List[dict]:
        items = self._call_collect(filename, "options")
        # options can be set multiple times - flatten
        flat: List[dict] = []
        for x in items:
            if isinstance(x, list):
                flat.extend(x)
        return flat
    
    def get_meta(self, filename: str) -> dict:
        items = self._call_collect(filename, "meta")
        meta: dict = {}
        if items:
            for x in items[0]:
                if isinstance(x, dict):
                    meta.update(x)
        return self._unescape(meta)
    
    def get_placeholders_with_options(self, filename: str) -> List[dict]:
        """
        Merged: placeholders (from file) + options (from Jinja blocks)
        - entries from options match by name
        - unknown placeholders get default: {"name":..., "type":"string"}
        """
        ph = self.get_placeholders(filename)
        opts = self.get_options(filename)
        
        opt_by_name = {o.get("name"): o for o in opts if isinstance(o, dict) and "name" in o}
        result: List[dict] = []
        
        for name in (ph - Reserved):
            if name in opt_by_name:
                # shallow copy is enough here
                entry = dict(opt_by_name[name])
                entry.setdefault("name", name)
                if entry.get("values"):
                    entry.setdefault("type", "enum")
                if not entry.get("type"):
                    entry.setdefault("type", "string")
                result.append(entry)
            else:
                result.append({"name": name, "type": "string"})
        return result
    
    def get_layout(self, filename: str):
        items = self._call_collect(filename, "layout")
        layout = []
        if items:
            for x in items[0]:
                if isinstance(x, dict):
                    for group, rows in x.items():
                        layout.append({"group": group, "rows": rows})
        return layout
    
    def list_templates(self, pattern: str = "*.docx") -> List[dict]:
        """
        Returns a list with templates + meta.
        Example: [{"filename": "example.docx", "meta":{}}]
        """
        files = sorted(p.name for p in self.base_dir.glob(pattern) if p.is_file())
        result = []
        for f in files:
            try:
                meta = self.get_meta(f)
                result.append({"filename": f, "meta": meta})
            except Exception:
                # Skip files that can't be read
                continue
        return result
    
    # ---------- rendering ----------
    def render_to_buffer(self, filename: str, context: Dict[str, Any]) -> BytesIO:
        if not isinstance(context, dict):
            raise InvalidContext("context must be dict")

        tpl = DocxTemplate(self._template_path(filename).as_posix())
        
        self._useReservedPlaceholders(context)
        
        jinja_env = Environment()
        jinja_env.globals["options"] = lambda *_args, **_kw: ""
        jinja_env.globals["meta"] = lambda *_args, **_kw: ""
        jinja_env.globals["layout"] = lambda *_args, **_kw: ""
        try:
            tpl.render(self._sanitize(context), jinja_env)
        except Exception as ex:
            raise RenderError(str(ex)) from ex
        buf = BytesIO()
        tpl.save(buf)
        buf.seek(0)
        return buf
    
    # ---------- read Excel ----------
    def _load_provider_addr(self, xlsx_name: str, provider_id: str) -> str:
        xlsx_path = self.base_dir / xlsx_name
        if not xlsx_path.exists():
            return ""
        
        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return ""
            
            headers = [str(h or "").strip().lower() for h in rows[0]]
            idx = {h: i for i, h in enumerate(headers)}
            
            for r in rows[1:]:
                if not r:
                    continue
                row = {h: str(r[idx[h]] or "").strip() for h in headers}
                if row.get("id") == provider_id:
                    return "\n".join([
                        row.get("name", ""),
                        row.get("street", ""),
                        f"{row.get('zip', '')} {row.get('city', '')}".strip()
                    ])
        except Exception:
            return ""
        
        return ""
